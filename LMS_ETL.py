import os
import traceback
from openpyxl import Workbook
from openpyxl import load_workbook
from win32com.client import Dispatch
from datetime import datetime, date
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formula.translate import Translator
wb0 = Workbook()
rgn_num = input('Are ICs included as a separate region?(Y/N) ').upper()
trn_name = input('What is the name of the training the report is for?(Formatted for File Name) ')
wb0.save('Statewide_Totals_'+ trn_name +'.xlsx')
if rgn_num == 'Y' or rgn_num == 'YES':
    rgn_num = 8
    region_list=['CCC', 'CNR', 'IC', 'NER', 'NWR', 'SCR', 'SER', 'SNR']
else:
    rgn_num = 7
    region_list=['CCC', 'CNR', 'NER', 'NWR', 'SCR', 'SER', 'SNR']
rgn_num = int(rgn_num)
x = 1
while x <= rgn_num:
    print("Making you some files you're likely going to ignore...")
    wb1 = load_workbook('data' + str(x) +'.xlsx')
    ws1 = wb1.active
    ws1.delete_cols(12,1)
    ws1.delete_cols(3,5)
    rn = region_list[x-1]
    ws1.title = rn + ' Totals'
    columns = ws1.max_column
    rows = ws1.max_row
    # Create fill
    redFill = PatternFill(start_color='ECBCBC', end_color='ECBCBC',fill_type='solid')
    red_text = Font(color="9C0006")
    # Add a conditional formatting based on a cell comparison
    # addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
    # Format if cell is less than 'formula'
    ws1.conditional_formatting.add('F2:F'+str(rows),CellIsRule(operator='lessThan', formula=[70], stopIfTrue=True, fill=redFill, font=red_text))
    wb1.save(rn + '.xlsx')
    x += 1


z = 1
while z <= rgn_num:
    try:
        path1 = 'C:\\Users\\garren-james\\LMS_ETL_temp\\Statewide_Totals_'+ trn_name +'.xlsx'
        path2 = 'C:\\Users\\garren-james\\LMS_ETL_temp\\'+ region_list[z-1] + '.xlsx'

        xl = Dispatch('Excel.Application')
        wb1 = xl.Workbooks.Open(Filename=path1)
        wb2 = xl.Workbooks.Open(Filename=path2)

        ws1 = wb2.Worksheets(1)
        ws1.Copy(Before=wb1.Worksheets(1))
        wb1.Close(SaveChanges=True)
        xl.Quit()

    except Exception as e:
        print('Unable to save report on shared system')
        print(str(e))
    finally:
        xl.Quit()
        z += 1


wb0 = load_workbook('Statewide_Totals_'+ trn_name +'.xlsx')
ws0 = wb0.get_sheet_by_name('Sheet')
ws0.title = 'Statewide Totals'
ws1 = wb0['CCC Totals']
for cell in ws1['A:A']:
    ws0.cell(row=cell.row, column=1, value=cell.value)


for cell in ws1['B:B']:
    ws0.cell(row=cell.row, column=2, value=cell.value)


ws0['C1'].value = ws1['C1'].value
ws0['D1'].value = ws1['D1'].value
ws0['E1'].value = ws1['E1'].value
ws0['F1'].value = ws1['F1'].value

if rgn_num == 'Y' or rgn_num == 'YES':
    ws0['C2'] = "=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'IC Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)"
    ws0['D2'] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'IC Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('D2')
    ws0['E2'] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'IC Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('E2')
    ws0['F2'] = "=C2/E2"
    ws0['F2'].style = 'Percent'

    columns = ws0.max_column
    rows = ws0.max_row
    #iterate formula down in column C
    c=3
    for row in ws0.iter_rows(min_col=3, min_row=3, max_col=3, max_row=rows):
        for cell in row:
            ws0['C'+str(c)] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'IC Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('C'+str(c))
            c+=1

    #iterate formula down in column D
    d=3
    for row in ws0.iter_rows(min_col=4, min_row=3, max_col=4, max_row=rows):
        for cell in row:
            ws0['D'+str(d)] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'IC Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('D'+str(d))
            d+=1


    #iterate formula down in column E
    e=3
    for row in ws0.iter_rows(min_col=5, min_row=3, max_col=5, max_row=rows):
        for cell in row:
            ws0['E'+str(e)] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'IC Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('E'+str(e))
            e+=1
else: 
    ws0['C2'] = "=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)"
    ws0['D2'] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2, 'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('D2')
    ws0['E2'] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('E2')
    ws0['F2'] = "=C2/E2"
    ws0['F2'].style = 'Percent'

    columns = ws0.max_column
    rows = ws0.max_row
    #iterate formula down in column C
    c=3
    for row in ws0.iter_rows(min_col=3, min_row=3, max_col=3, max_row=rows):
        for cell in row:
            print('Doing your work!')
            ws0['C'+str(c)] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2, 'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('C'+str(c))
            c+=1

    #iterate formula down in column D
    d=3
    for row in ws0.iter_rows(min_col=4, min_row=3, max_col=4, max_row=rows):
        for cell in row:
            print('Doing some more of your work!')
            ws0['D'+str(d)] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2, 'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('D'+str(d))
            d+=1


    #iterate formula down in column E
    e=3
    for row in ws0.iter_rows(min_col=5, min_row=3, max_col=5, max_row=rows):
        for cell in row:
            print('Still working on it!')
            ws0['E'+str(e)] = Translator("=SUM('SNR Totals'!C2,'SER Totals'!C2,'SCR Totals'!C2,'NWR Totals'!C2,'NER Totals'!C2,'CNR Totals'!C2,'CCC Totals'!C2)", 'C2').translate_formula('E'+str(e))
            e+=1

#iterate formula down in column f
f=3
for row in ws0.iter_rows(min_col=5, min_row=3, max_col=5, max_row=rows):
    for cell in row:
        ws0['F'+str(f)] = '=C' + str(f) + '/E' + str(f)
        ws0['F'+str(f)].style = 'Percent'
        f+=1


redFill = PatternFill(start_color='ECBCBC', end_color='ECBCBC',fill_type='solid')
red_text = Font(color="9C0006")
# Add a conditional formatting based on a cell comparison
# addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
# Format if cell is less than 'formula'
ws0.conditional_formatting.add('F2:F'+str(rows),CellIsRule(operator='lessThan', formula=[.70], stopIfTrue=True, fill=redFill, font=red_text))

print("All finished with Statewide_Totals_"+ trn_name +".xlsx  Now it's your turn to do something!")
wb0.save('Statewide_Totals_'+ trn_name +'.xlsx')