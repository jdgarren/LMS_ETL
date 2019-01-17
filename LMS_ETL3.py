import os
import traceback
from openpyxl import Workbook
from openpyxl import load_workbook
from win32com.client import Dispatch
from datetime import datetime, date
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
wb0 = Workbook()
wb0.save('Statewide_Totals.xlsx')
rgn_num = input('Are ICs included as a separate region?(Y/N) ').upper()
if rgn_num == 'Y' or rgn_num == 'YES':
    rgn_num = 8
    region_list=['CCC', 'CNR', 'IC', 'NER', 'NWR', 'SCR', 'SER', 'SNR']
else:
    rgn_num = 7
    region_list=['CCC', 'CNR', 'NER', 'NWR', 'SCR', 'SER', 'SNR']
rgn_num = int(rgn_num)
x = 1
while x <= rgn_num:
    wb1 = load_workbook('data' + str(x) +'.xlsx')
    ws1 = wb1.active
    ws1.delete_cols(12,1)
    ws1.delete_cols(3,5)
    rn = region_list[x-1]
    ws1.title = rn + ' Totals'
    columns = ws1.max_column
    rows = ws1.max_row
    ws1.cell(row=1, column=columns+1, value='Region')
    for row in ws1.iter_rows(min_col=columns+1, min_row=2, max_col=columns+1, max_row=rows):
        for cell in row:
            cell.value=rn

    # Create fill
    redFill = PatternFill(start_color='ECBCBC', end_color='ECBCBC',fill_type='solid')
    red_text = Font(color="9C0006")
    # Add a conditional formatting based on a cell comparison
    # addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
    # Format if cell is less than 'formula'
    ws1.conditional_formatting.add('F2:F'+str(rows),CellIsRule(operator='lessThan', formula=[70], stopIfTrue=True, fill=redFill, font=red_text))
    wb1.save(rn + '.xlsx')
    x += 1


z = 1
while z <= rgn_num:
    try:
        path1 = 'C:\\Users\\garren-james\\Python Projects\\Training Data\\Statewide_Totals.xlsx'
        path2 = 'C:\\Users\\garren-james\\Python Projects\\Training Data\\'+ region_list[z-1] + '.xlsx'

        xl = Dispatch('Excel.Application')
        wb1 = xl.Workbooks.Open(Filename=path1)
        wb2 = xl.Workbooks.Open(Filename=path2)

        ws1 = wb2.Worksheets(1)
        ws1.Copy(Before=wb1.Worksheets(1))
        wb1.Close(SaveChanges=True)
        xl.Quit()

    except Exception as e:
        print('Unable to save report on shared system')
        print(str(e))
    finally:
        xl.Quit()
        z += 1
